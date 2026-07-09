import csv
import io
import datetime
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

class ExportService:
    HEADERS = [
        "Request Number",
        "Parent Name",
        "Child Name",
        "Relationship",
        "Mobile Number",
        "Email",
        "Date of Birth",
        "Gender",
        "Appointment Type",
        "Primary Concern",
        "Preferred Date",
        "Preferred Slot",
        "Booking Source",
        "Status",
        "Created At"
    ]

    @classmethod
    def get_row_data(cls, req) -> list:
        return [
            req.request_number,
            f"{req.parent_first_name} {req.parent_last_name}".strip(),
            f"{req.child_first_name} {req.child_last_name}".strip(),
            req.get_relationship_to_child_display(),
            req.mobile_number,
            req.email,
            req.date_of_birth.strftime('%Y-%m-%d') if req.date_of_birth else "N/A",
            req.get_gender_display(),
            req.get_appointment_type_display(),
            req.primary_concern,
            req.preferred_date.strftime('%Y-%m-%d') if req.preferred_date else "N/A",
            req.preferred_time_slot,
            req.get_booking_source_display(),
            req.get_status_display(),
            req.created_at.strftime('%Y-%m-%d %H:%M:%S') if req.created_at else "N/A"
        ]

    @classmethod
    def export_csv(cls, queryset) -> bytes:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(cls.HEADERS)
        for req in queryset:
            writer.writerow(cls.get_row_data(req))
        return output.getvalue().encode('utf-8')

    @classmethod
    def export_excel(cls, queryset) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Appointment Requests"

        # Styles
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
        alignment_center = Alignment(horizontal="center", vertical="center")
        alignment_left = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # Header Row
        ws.append(cls.HEADERS)
        for col_num in range(1, len(cls.HEADERS) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 28

        # Data Rows
        row_num = 2
        for req in queryset:
            row_data = cls.get_row_data(req)
            ws.append(row_data)
            for col_num in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = alignment_left
                cell.border = thin_border
            ws.row_dimensions[row_num].height = 20
            row_num += 1

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        return excel_bytes

    @classmethod
    def export_pdf(cls, queryset) -> bytes:
        # Generates a landscape table list PDF of filtered appointment requests
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=18,
            leading=22,
            textColor=colors.HexColor('#0F172A'),
            spaceAfter=15
        )

        hdr_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            leading=10,
            textColor=colors.HexColor('#FFFFFF')
        )

        body_style = ParagraphStyle(
            'BodyStyle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=7,
            leading=9,
            textColor=colors.HexColor('#0F172A')
        )

        story = []
        story.append(Paragraph("Appointment Requests Export List", title_style))
        story.append(Spacer(1, 10))

        # We will show subset of fields in PDF to avoid horizontal overflow
        pdf_headers = ["Req Number", "Child Name", "Parent Name", "Mobile", "Preferred Date", "Type", "Source", "Status"]
        table_data = [[Paragraph(h, hdr_style) for h in pdf_headers]]

        for req in queryset:
            table_data.append([
                Paragraph(req.request_number, body_style),
                Paragraph(f"{req.child_first_name} {req.child_last_name}".strip(), body_style),
                Paragraph(f"{req.parent_first_name} {req.parent_last_name}".strip(), body_style),
                Paragraph(req.mobile_number, body_style),
                Paragraph(req.preferred_date.strftime('%Y-%m-%d') if req.preferred_date else "N/A", body_style),
                Paragraph(req.get_appointment_type_display(), body_style),
                Paragraph(req.get_booking_source_display(), body_style),
                Paragraph(req.get_status_display(), body_style)
            ])

        # Page width in landscape is 792pt. Margins are 36pt left and 36pt right. Usable width is 720pt.
        col_widths = [80, 100, 100, 75, 75, 100, 90, 100]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A8A')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#FFFFFF'), colors.HexColor('#F8FAFC')]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0'))
        ]))
        story.append(t)
        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes
